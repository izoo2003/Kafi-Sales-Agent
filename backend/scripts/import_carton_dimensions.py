"""Import CARTONS SIZE LIST.xlsx into backend/data/kafi_carton_dimensions.json."""

import json
import re
import sys
from pathlib import Path

import openpyxl

BACKEND_DIR = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = Path.home() / "Downloads" / "CARTONS SIZE LIST (1).xlsx"
OUT = BACKEND_DIR / "data" / "kafi_carton_dimensions.json"

_DIM_SEP = re.compile(r"\s*[xX×]\s*|\s+", re.I)


def _parse_dim_string(raw: str | None) -> tuple[float | None, float | None, float | None]:
    """Parse '18 X 11 X 9.5' or '12 X 12 11.5' into L/W/H in cm."""
    if not raw or not str(raw).strip():
        return None, None, None
    text = str(raw).strip().replace("�", "")
    if text.lower() in {"z`", "z'", "z"}:
        return None, None, None

    parts = [p for p in _DIM_SEP.split(text) if p and re.match(r"^[\d.]+$", p)]
    if len(parts) >= 3:
        return float(parts[0]), float(parts[1]), float(parts[2])
    if len(parts) == 2:
        return float(parts[0]), float(parts[1]), None
    return None, None, None


def _num(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except ValueError:
        return None


def _int_num(value) -> int | None:
    n = _num(value)
    return int(n) if n is not None else None


def _entry(
    *,
    sno: int,
    product: str,
    packing: str | None,
    length_cm: float | None,
    width_cm: float | None,
    height_cm: float | None,
    cbm: float | None,
    fcl_20: int | None,
    fcl_40: int | None,
    fcl_40hc: int | None,
    sheet: str,
    brand: str | None = None,
) -> dict:
    if length_cm and width_cm and height_cm:
        cbm_calc = round((length_cm * width_cm * height_cm) / 1_000_000, 4)
    else:
        cbm_calc = None
    return {
        "id": sno,
        "product": product.strip(),
        "packing": (packing or "").strip() or None,
        "brand": brand,
        "length_cm": length_cm,
        "width_cm": width_cm,
        "height_cm": height_cm,
        "cbm": cbm if cbm is not None else cbm_calc,
        "fcl_20ft_cartons": fcl_20,
        "fcl_40ft_cartons": fcl_40,
        "fcl_40hc_cartons": fcl_40hc,
        "source_sheet": sheet,
    }


def _brand_from_product(name: str) -> str | None:
    upper = name.upper()
    if "ESSENCE" in upper or "ESSNCE" in upper:
        return "ESSENCE"
    if "AL SHIFA" in upper or "MUJEZAT" in upper:
        return "AL SHIFA"
    return None


def import_old_sheet(ws) -> list[dict]:
    items: list[dict] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        cells = list(row) + [None] * 12
        sno = cells[1]
        desc = cells[2]
        if not isinstance(sno, (int, float)) or not desc or not str(desc).strip():
            continue
        if str(desc).strip().upper() in {"DESCRIPTION ", "S.NO"}:
            continue

        l, w, h = _parse_dim_string(cells[3])
        items.append(
            _entry(
                sno=int(sno),
                product=str(desc).strip(),
                packing=None,
                length_cm=l,
                width_cm=w,
                height_cm=h,
                cbm=_num(cells[4]),
                fcl_20=_int_num(cells[6]),
                fcl_40=_int_num(cells[7]),
                fcl_40hc=_int_num(cells[8]),
                sheet="OLD",
                brand=_brand_from_product(str(desc)),
            )
        )
    return items


def import_new_sheet(ws) -> list[dict]:
    items: list[dict] = []
    current_brand_group: str | None = None

    for row in ws.iter_rows(min_row=1, values_only=True):
        cells = list(row) + [None] * 15
        label = str(cells[2] or "").strip().upper()
        if label.endswith("BRAND ") or label.endswith("BRAND"):
            current_brand_group = str(cells[2]).strip()
            continue

        sno = cells[1]
        product = cells[2]
        if not isinstance(sno, (int, float)) or not product or not str(product).strip():
            continue
        if str(product).strip().upper() in {"PRODUCT ", "S.NO", "DESCRIPTION "}:
            continue

        product_str = str(product).strip()
        brand = _brand_from_product(product_str) or (
            "ESSENCE" if current_brand_group and "ESSENCE" in current_brand_group.upper() else None
        )
        if not brand and current_brand_group:
            brand = current_brand_group.replace(" BRAND", "").strip() or None

        items.append(
            _entry(
                sno=int(sno),
                product=product_str,
                packing=str(cells[3]).strip() if cells[3] else None,
                length_cm=_num(cells[4]),
                width_cm=_num(cells[5]),
                height_cm=_num(cells[6]),
                cbm=None,
                fcl_20=_int_num(cells[7]),
                fcl_40=_int_num(cells[8]),
                fcl_40hc=_int_num(cells[9]),
                sheet="NEW",
                brand=brand,
            )
        )
    return items


def import_workbook(source: Path) -> dict:
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    old_items = import_old_sheet(wb["OLD "]) if "OLD " in wb.sheetnames else []
    new_items = import_new_sheet(wb["NEW"]) if "NEW" in wb.sheetnames else []
    wb.close()

    # Prefer NEW sheet entries; keep OLD-only products
    new_keys = {(i["product"].lower(), (i["packing"] or "").lower()) for i in new_items}
    merged = list(new_items)
    next_id = len(merged) + 1
    for item in old_items:
        key = (item["product"].lower(), (item["packing"] or "").lower())
        if key not in new_keys:
            item = {**item, "id": next_id}
            merged.append(item)
            next_id += 1

    essence = [i for i in merged if i.get("brand") == "ESSENCE"]
    return {
        "source": source.name,
        "company": "Kafi Commodities (Pvt.) Limited",
        "entry_count": len(merged),
        "essence_count": len(essence),
        "sheets": {"OLD": len(old_items), "NEW": len(new_items)},
        "entries": merged,
    }


def main() -> None:
    source = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SOURCE
    if not source.exists():
        raise SystemExit(f"Source file not found: {source}")

    data = import_workbook(source)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {data['entry_count']} carton entries ({data['essence_count']} ESSENCE) to {OUT}")


if __name__ == "__main__":
    main()
