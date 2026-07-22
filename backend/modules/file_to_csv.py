"""Convert uploaded spreadsheet files to CSV for lead import."""

from __future__ import annotations

import csv
import io
from pathlib import Path

from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

SUPPORTED_UPLOAD_EXTENSIONS = {".csv", ".xlsx", ".xls", ".xlsm", ".tsv"}


def _sniff_extension(raw: bytes) -> str:
    if raw[:2] == b"PK":
        return ".xlsx"
    if raw[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return ".xls"
    head = raw[:4096]
    try:
        text = head.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = head.decode("latin-1")
        except UnicodeDecodeError:
            text = ""
    lowered = text.lower()
    if "<table" in lowered or "urn:schemas-microsoft-com:office:excel" in lowered:
        return ".xls"
    if text.count("\t") > max(text.count(","), text.count(";")) and text.count("\t") > 2:
        return ".tsv"
    if text.count(";") > text.count(",") and text.count(";") > 2:
        return ".csv"
    return ".csv"


def _resolve_extension(filename: str | None, raw: bytes) -> str:
    ext = Path((filename or "").strip()).suffix.lower()
    if ext in SUPPORTED_UPLOAD_EXTENSIONS:
        return ext
    return _sniff_extension(raw)


def _cell_text(cell) -> str:
    link = cell.find("a")
    if link:
        href = (link.get("href") or "").strip()
        if href.startswith(("http://", "https://", "mailto:")):
            if href.startswith("mailto:"):
                return href.removeprefix("mailto:").strip()
            return href
        link_text = link.get_text(" ", strip=True)
        if link_text:
            return link_text
    return cell.get_text(" ", strip=True)


def _html_table_to_csv(html: str) -> str:
    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table")
    if not table:
        raise ValueError("No table found in this Excel file.")

    out = io.StringIO()
    writer = csv.writer(out)
    for row in table.find_all("tr"):
        cells = [_cell_text(cell) for cell in row.find_all(["th", "td"])]
        if any(cells):
            writer.writerow(cells)
    content = out.getvalue()
    if not content.strip():
        raise ValueError("The uploaded file has no data rows.")
    return content


def _spreadsheet_rows_to_csv(rows) -> str:
    out = io.StringIO()
    writer = csv.writer(out)
    wrote_row = False
    for row in rows:
        cells = [
            ""
            if value is None
            else str(value).replace("\r", " ").replace("\n", " ").strip()
            for value in row
        ]
        if not any(cell.strip() for cell in cells):
            continue
        writer.writerow(cells)
        wrote_row = True
    content = out.getvalue()
    if not wrote_row:
        raise ValueError("The uploaded spreadsheet has no data rows.")
    return content


def _xlsx_to_csv(raw: bytes) -> str:
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except InvalidFileException as exc:
        raise ValueError(
            "Could not read this Excel file. Save it as .xlsx or .csv and upload again."
        ) from exc
    try:
        sheet = workbook.active
        # Stream cells → CSV without buffering the whole sheet in a list first.
        return _spreadsheet_rows_to_csv(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()


def _binary_xls_to_csv(raw: bytes) -> str:
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError(
            "Binary .xls support requires xlrd. Run: pip install xlrd — or save the file as .xlsx / .csv."
        ) from exc
    try:
        book = xlrd.open_workbook(file_contents=raw)
    except xlrd.XLRDError as exc:
        raise ValueError(
            "Could not read this .xls file. Save it as .xlsx or .csv and upload again."
        ) from exc
    sheet = book.sheet_by_index(0)

    def _iter_rows():
        for row_idx in range(sheet.nrows):
            yield [sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)]

    return _spreadsheet_rows_to_csv(_iter_rows())


def _delimiter_separated_to_csv(content: str, delimiter: str) -> str:
    reader = csv.reader(io.StringIO(content), delimiter=delimiter)
    out = io.StringIO()
    writer = csv.writer(out)
    wrote_row = False
    for row in reader:
        if not row or not any(cell.strip() for cell in row):
            continue
        writer.writerow([cell.strip() for cell in row])
        wrote_row = True
    csv_content = out.getvalue()
    if not wrote_row:
        raise ValueError("The uploaded file has no data rows.")
    return csv_content


def convert_upload_to_csv(filename: str | None, raw: bytes) -> tuple[str, list[str]]:
    if not raw:
        raise ValueError("The uploaded file is empty.")

    ext = _resolve_extension(filename, raw)
    display_name = (filename or "upload").strip() or "upload"
    messages: list[str] = []

    if ext == ".csv":
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw.decode("latin-1")
        try:
            dialect = csv.Sniffer().sniff(text[:8192], delimiters=",\t;|")
            if dialect.delimiter != ",":
                messages.append(f"Detected {display_name!r} as delimited text and converted to CSV.")
                return _delimiter_separated_to_csv(text, dialect.delimiter), messages
        except csv.Error:
            pass
        return text, messages

    if ext == ".tsv":
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw.decode("latin-1")
        messages.append(f"Converted {display_name} from TSV to CSV.")
        return _delimiter_separated_to_csv(text, "\t"), messages

    if ext in {".xlsx", ".xlsm"}:
        messages.append(f"Converted {display_name} from Excel ({ext}) to CSV.")
        return _xlsx_to_csv(raw), messages

    if ext == ".xls":
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = None
        if text and "<table" in text.lower():
            messages.append(f"Converted {display_name} from Excel (.xls) to CSV.")
            return _html_table_to_csv(text), messages
        messages.append(f"Converted {display_name} from Excel (.xls) to CSV.")
        return _binary_xls_to_csv(raw), messages

    supported = ", ".join(sorted(SUPPORTED_UPLOAD_EXTENSIONS))
    raise ValueError(f"Unsupported file type. Upload one of: {supported}")
